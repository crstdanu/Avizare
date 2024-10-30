from openpyxl import load_workbook


wb = load_workbook("G:/Shared drives/Root/11. DATABASE/01. Automatizari avize/Executie/01. Pentru incepere/Grafic executie/Model Grafic executie.xlsx")

ws = wb.active


inceput = "luna 1"
final = "luna 2"
ws['C10'] = inceput
ws['D10'] = final

if inceput == final:
    pass


ws.merge_cells('C10:D10')
ws['C11'] = 'Aici pun data de inceput'
ws['D11'] = 'Aici pun data de final'


wb.save("G:/Shared drives/Root/11. DATABASE/01. Automatizari avize/Executie/01. Pentru incepere/Grafic executie/Output.xlsx")