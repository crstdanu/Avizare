import pyodbc
import os
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
from datetime import datetime as dt
import win32com.client as win32
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
import PyPDF2
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
from dateutil.relativedelta import relativedelta
import math
import time


pagina_goala = os.path.join(os.path.dirname(__file__), 'DOCUMENTE', 'pagina_goala.pdf')


def xlsx_to_pdf(excel_file):
    excel = win32.Dispatch('Excel.Application')
    workbook = excel.Workbooks.Open(excel_file)
    excel.Visible = False
    pdf_file = excel_file.replace(".xlsx", r".pdf")
    workbook.ExportAsFixedFormat(0, pdf_file)
    workbook.Close(False)
    excel.Quit()
    return pdf_file


def count_pages(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        num_pages = len(reader.pages)
    return num_pages


def convert_to_pdf(doc):
    word = win32.DispatchEx("Word.Application")
    new_name = doc.replace(".docx", r".pdf")
    try:
        word.Visible = False
        worddoc = word.Documents.Open(doc)
        worddoc.SaveAs(new_name, FileFormat=17)
        worddoc.Close()
        return new_name
    finally:
        # Ensure Word process is terminated to avoid file locks
        try:
            word.Quit()
        except Exception:
            pass


def get_today_date():
    today = dt.today().date()
    return today.strftime("%d-%m-%Y")


def get_db_connection():
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=G:/Shared drives/Root/11. DATABASE/RoGoTehnic-DataBase-2024.accdb;"
    return pyodbc.connect(con_string)


def fetch_single_value(cursor, query, params):
    return cursor.execute(query, params).fetchval()


def get_Firma_proiectare(cursor, id_firma_proiectare):
    return {
        'nume': fetch_single_value(cursor, 'SELECT NumeFirma FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'reprezentant': fetch_single_value(cursor, 'SELECT Nume FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'localitate_repr': fetch_single_value(cursor, 'SELECT Localitate FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'adresa_repr': fetch_single_value(cursor, 'SELECT Adresa FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'judet_repr': fetch_single_value(cursor, 'SELECT Judet FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'seria_CI': fetch_single_value(cursor, 'SELECT SerieCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'nr_CI': fetch_single_value(cursor, 'SELECT NumarCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'data_CI': fetch_single_value(cursor, 'SELECT DataCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'cnp_repr': fetch_single_value(cursor, 'SELECT CNP FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'caleCI': fetch_single_value(cursor, 'SELECT CaleCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeProiectare WHERE IDFirma = ?)', (id_firma_proiectare,)),
        'CaleStampila': fetch_single_value(cursor, 'SELECT CaleStampila FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
        'CaleCertificat': fetch_single_value(cursor, 'SELECT CaleCertificat FROM tblFirmeProiectare WHERE IDFirma = ?', (id_firma_proiectare,)),
    }


def get_Firma_executie(cursor, id_firma_executie):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate FROM tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'reprezentant': fetch_single_value(cursor, 'SELECT Nume FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'localitate_repr': fetch_single_value(cursor, 'SELECT NumeLocalitate FROM tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?))', (id_firma_executie,)),
        'adresa_repr': fetch_single_value(cursor, 'SELECT Adresa FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'judet_repr': fetch_single_value(cursor, 'SELECT Judet FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'seria_CI': fetch_single_value(cursor, 'SELECT SerieCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'nr_CI': fetch_single_value(cursor, 'SELECT NumarCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'data_CI': fetch_single_value(cursor, 'SELECT DataCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'cnp_repr': fetch_single_value(cursor, 'SELECT CNP FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'CaleCI': fetch_single_value(cursor, 'SELECT CaleCI FROM tblAngajati WHERE IDAngajat = (SELECT IDReprezentant FROM tblFirmeExecutie WHERE IDFirma = ?)', (id_firma_executie,)),
        'CaleStampila': fetch_single_value(cursor, 'SELECT CaleStampila FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
        'CaleCertificat': fetch_single_value(cursor, 'SELECT CaleCertificat FROM tblFirmeExecutie WHERE IDFirma = ?', (id_firma_executie,)),
    }


def get_Client(cursor, id_client):
    return {
        'nume': fetch_single_value(cursor, 'SELECT NumeClient FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblClienti WHERE IDClient = ?)', (id_client,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblClienti WHERE IDClient = ?)', (id_client,)),
        'tip_client': fetch_single_value(cursor, 'SELECT TipClient FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'reprezentant': fetch_single_value(cursor, 'SELECT Reprezentant FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblClienti WHERE IDClient = ?', (id_client,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblClienti WHERE IDClient = ?', (id_client,)),
    }


def get_CU(cursor, id_lucrare):
    return {
        'EmitentCU': fetch_single_value(cursor, 'SELECT EmitentCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'NrCU': fetch_single_value(cursor, 'SELECT NrCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataCU': fetch_single_value(cursor, 'SELECT DataCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'FacturarePeClient': fetch_single_value(cursor, 'SELECT FacturarePeClient FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DescriereaProiectului': fetch_single_value(cursor, 'SELECT DescriereaProiectului FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'IDIntocmit': fetch_single_value(cursor, 'SELECT IDIntocmit FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'IDVerificat': fetch_single_value(cursor, 'SELECT IDVerificat FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleChitantaAPM': fetch_single_value(cursor, 'SELECT CaleChitantaAPM FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleCU': fetch_single_value(cursor, 'SELECT CaleCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanIncadrareCU': fetch_single_value(cursor, 'SELECT CalePlanIncadrareCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatieCU': fetch_single_value(cursor, 'SELECT CalePlanSituatieCU FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleMemoriuTehnicSS': fetch_single_value(cursor, 'SELECT CaleMemoriuTehnicSS FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleActeBeneficiar': fetch_single_value(cursor, 'SELECT CaleActeBeneficiar FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleActeFacturare': fetch_single_value(cursor, 'SELECT CaleActeFacturare FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatieDWG': fetch_single_value(cursor, 'SELECT CalePlanSituatieDWG FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatiePDF': fetch_single_value(cursor, 'SELECT CalePlanSituatiePDF FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleACCUConstructie': fetch_single_value(cursor, 'SELECT CaleACCUConstructie FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleExtraseCF': fetch_single_value(cursor, 'SELECT CaleExtraseCF FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleAvizGiS': fetch_single_value(cursor, 'SELECT CaleAvizGiS FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleRidicareTopoDWG': fetch_single_value(cursor, 'SELECT CaleRidicareTopoDWG FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'SuprafataOcupata': fetch_single_value(cursor, 'SELECT SuprafataOcupata FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'LungimeTraseu': fetch_single_value(cursor, 'SELECT LungimeTraseu FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleATR': fetch_single_value(cursor, 'SELECT CaleATR FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleAvizCTE': fetch_single_value(cursor, 'SELECT CaleAvizCTE FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleChitantaDSP': fetch_single_value(cursor, 'SELECT CaleChitantaDSP FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleChitantaPolitie': fetch_single_value(cursor, 'SELECT CaleChitantaPolitie FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleChitantaPolitie': fetch_single_value(cursor, 'SELECT CaleChitantaPolitie FROM tblCU WHERE ID_Lucrare = ?', (id_lucrare,)),
    }


def get_IncepereExecutie(cursor, id_lucrare):
    return {
        'NumarAC': fetch_single_value(cursor, 'SELECT NumarAC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataAC': fetch_single_value(cursor, 'SELECT DataAC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ValoareAC': fetch_single_value(cursor, 'SELECT ValoareAC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDovadaPlataAC': fetch_single_value(cursor, 'SELECT CaleDovadaPlataAC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDovadaPlataISC': fetch_single_value(cursor, 'SELECT CaleDovadaPlataISC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ValabilitateAC': fetch_single_value(cursor, 'SELECT ValabilitateAC FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ValabilitateExecutie': fetch_single_value(cursor, 'SELECT ValabilitateExecutie FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataIncepereExecutie': fetch_single_value(cursor, 'SELECT DataIncepereExecutie FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'NrCL': fetch_single_value(cursor, 'SELECT NrCL FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'NrDeciziePersonal': fetch_single_value(cursor, 'SELECT NrDeciziePersonal FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'GraficExecutie': fetch_single_value(cursor, 'SELECT GraficExecutie FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataIncepereGrafic': fetch_single_value(cursor, 'SELECT DataIncepereGrafic FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataFinalizareGrafic': fetch_single_value(cursor, 'SELECT DataFinalizareGrafic FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleContractCitadin': fetch_single_value(cursor, 'SELECT CaleContractCitadin FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleContractSpatiiVerzi': fetch_single_value(cursor, 'SELECT CaleContractSpatiiVerzi FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleContractForaj': fetch_single_value(cursor, 'SELECT CaleContractForaj FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPSapareSant': fetch_single_value(cursor, 'SELECT DPSapareSant FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPPozareCablu': fetch_single_value(cursor, 'SELECT DPPozareCablu FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPMontarePTAV': fetch_single_value(cursor, 'SELECT DPMontarePTAV FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPMontareStalpi': fetch_single_value(cursor, 'SELECT DPMontareStalpi FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPExecutieForaj': fetch_single_value(cursor, 'SELECT DPExecutieForaj FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPMontareFiride': fetch_single_value(cursor, 'SELECT DPMontareFiride FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DPStareInitiala': fetch_single_value(cursor, 'SELECT DPStareInitiala FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleACScanat': fetch_single_value(cursor, 'SELECT CaleACScanat FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanIncadrareACScanat': fetch_single_value(cursor, 'SELECT CalePlanIncadrareACScanat FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatieACScanat': fetch_single_value(cursor, 'SELECT CalePlanSituatieACScanat FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleMemoriuTehnicACScanat': fetch_single_value(cursor, 'SELECT CaleMemoriuTehnicACScanat FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanIncadrarePTH': fetch_single_value(cursor, 'SELECT CalePlanIncadrarePTH FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatiePTH': fetch_single_value(cursor, 'SELECT CalePlanSituatiePTH FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleSchemaMonofilaraJT': fetch_single_value(cursor, 'SELECT CaleSchemaMonofilaraJT FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleSchemaMonofilaraMT': fetch_single_value(cursor, 'SELECT CaleSchemaMonofilaraMT FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleInstruireColectiva': fetch_single_value(cursor, 'SELECT CaleInstruireColectiva FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ContractExecutieDELGAZ': fetch_single_value(cursor, 'SELECT ContractExecutieDELGAZ FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleContractRacordare': fetch_single_value(cursor, 'SELECT CaleContractRacordare FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleContractExecutie': fetch_single_value(cursor, 'SELECT CaleContractExecutie FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ID_DiriginteSantier': fetch_single_value(cursor, 'SELECT ID_DiriginteSantier FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ID_RTE': fetch_single_value(cursor, 'SELECT ID_RTE FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ID_RTEConstructii': fetch_single_value(cursor, 'SELECT ID_RTEConstructii FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ID_SSM': fetch_single_value(cursor, 'SELECT ID_SSM FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
        'ID_ManagerProiect': fetch_single_value(cursor, 'SELECT ID_ManagerProiect FROM tblIncepereExecutie WHERE ID_Lucrare = ?', (id_lucrare,)),
    }


def get_Finalizare(cursor, id_lucrare):
    return {
        'ValoareReala': fetch_single_value(cursor, 'SELECT ValoareReala FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'DataFinalizare': fetch_single_value(cursor, 'SELECT DataFinalizare FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleImputernicireDelgaz': fetch_single_value(cursor, 'SELECT CaleImputernicireDelgaz FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleAnuntIncepereUAT': fetch_single_value(cursor, 'SELECT CaleAnuntIncepereUAT FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDevizFinal': fetch_single_value(cursor, 'SELECT CaleDevizFinal FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleFacturiRGT': fetch_single_value(cursor, 'SELECT CaleFacturiRGT FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDovadaPlataFacturi': fetch_single_value(cursor, 'SELECT CaleDovadaPlataFacturi FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleReferatDS': fetch_single_value(cursor, 'SELECT CaleReferatDS FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleRaportProiectant': fetch_single_value(cursor, 'SELECT CaleRaportProiectant FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleITL': fetch_single_value(cursor, 'SELECT CaleITL FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleAnuntFinalizareUAT': fetch_single_value(cursor, 'SELECT CaleAnuntFinalizareUAT FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDovadaRegularizareTaxaISC': fetch_single_value(cursor, 'SELECT CaleDovadaRegularizareTaxaISC FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDovadaRegularizareTaxaAC': fetch_single_value(cursor, 'SELECT CaleDovadaRegularizareTaxaAC FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleAdeverintaISC': fetch_single_value(cursor, 'SELECT CaleAdeverintaISC FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleDispozitieSantier': fetch_single_value(cursor, 'SELECT CaleDispozitieSantier FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanIncadrareDS': fetch_single_value(cursor, 'SELECT CalePlanIncadrareDS FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CalePlanSituatieDS': fetch_single_value(cursor, 'SELECT CalePlanSituatieDS FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
        'CaleRaspunsUatDS': fetch_single_value(cursor, 'SELECT CaleRaspunsUatDS FROM tblFinalizare WHERE ID_Lucrare = ?', (id_lucrare,)),
    }


def get_EmitentCU(cursor, IDEmitent):
    return {
        'denumire_institutie': fetch_single_value(cursor, 'SELECT DenumireInstitutie FROM tblEmitentiCU WHERE IDEmitent = ?', (IDEmitent,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT LocalitateInstitutie FROM tblEmitentiCU WHERE IDEmitent = ?)', (IDEmitent,)),
        'adresa': fetch_single_value(cursor, 'SELECT AdresaInstitutie FROM tblEmitentiCU WHERE IDEmitent = ?', (IDEmitent,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT JudetInstitutie FROM tblEmitentiCU WHERE IDEmitent = ?)', (IDEmitent,)),
        'cod_postal': fetch_single_value(cursor, 'SELECT CodPostal FROM tblEmitentiCU WHERE IDEmitent = ?', (IDEmitent,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblEmitentiCU WHERE IDEmitent = ?', (IDEmitent,)),
        'email': fetch_single_value(cursor, 'SELECT AdresaEmail FROM tblEmitentiCU WHERE IDEmitent = ?', (IDEmitent,)),
    }


def get_Diriginte_Santier(cursor, IDDiriginteSantier):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblMP_DS WHERE IDMPDS = ?)', (IDDiriginteSantier,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblMP_DS WHERE IDMPDS = ?)', (IDDiriginteSantier,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblMP_DS WHERE IDMPDS = ?', (IDDiriginteSantier,)),
    }


def get_Manager_Proiect(cursor, ID_ManagerProiect):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblMP_DS WHERE IDMPDS = ?)', (ID_ManagerProiect,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblMP_DS WHERE IDMPDS = ?)', (ID_ManagerProiect,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblMP_DS WHERE IDMPDS = ?', (ID_ManagerProiect,)),
    }


def get_RTE(cursor, IDRTEElectric):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblRTE WHERE IDRTE = ?)', (IDRTEElectric,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblRTE WHERE IDRTE = ?)', (IDRTEElectric,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblRTE WHERE IDRTE = ?', (IDRTEElectric,)),
    }


def get_RTE_constructii(cursor, IDRTEConstructii):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblRTEConstructii WHERE IDResponsabil = ?', (IDRTEConstructii,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblRTEConstructii WHERE IDResponsabil = ?)', (IDRTEConstructii,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblRTEConstructii WHERE IDResponsabil = ?', (IDRTEConstructii,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblRTEConstructii WHERE IDResponsabil = ?)', (IDRTEConstructii,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblRTEConstructii WHERE IDResponsabil = ?', (IDRTEConstructii,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblRTEConstructii WHERE IDResponsabil = ?', (IDRTEConstructii,)),
    }


def get_responsabil_SSM(cursor, IDResponsabilSSM):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblSSM WHERE ID = ?', (IDResponsabilSSM,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblSSM WHERE ID = ?)', (IDResponsabilSSM,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblSSM WHERE ID = ?', (IDResponsabilSSM,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblSSM WHERE ID = ?)', (IDResponsabilSSM,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblSSM WHERE ID = ?', (IDResponsabilSSM,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblSSM WHERE ID = ?', (IDResponsabilSSM,)),
    }


def get_Beneficiar(cursor, id_beneficiar):
    return {
        'nume': fetch_single_value(cursor, 'SELECT NumeBeneficiar FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT Localitate FROM tblBeneficiari WHERE IDBeneficiar = ?)', (id_beneficiar,)),
        'adresa': fetch_single_value(cursor, 'SELECT Adresa FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet from tblJudete where ID_Judet = (SELECT Judet FROM tblBeneficiari WHERE IDBeneficiar = ?)', (id_beneficiar,)),
        'CodPostal': fetch_single_value(cursor, 'SELECT CodPostal FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'CUI': fetch_single_value(cursor, 'SELECT CUI FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'NrRegCom': fetch_single_value(cursor, 'SELECT NrRegCom FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'reprezentant': fetch_single_value(cursor, 'SELECT Reprezentant FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
        'email': fetch_single_value(cursor, 'SELECT Email FROM tblBeneficiari WHERE IDBeneficiar = ?', (id_beneficiar,)),
    }


def get_Lucrare(cursor, id_lucrare):
    return {
        'nume': fetch_single_value(cursor, 'SELECT DenumireLucrare FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'localitate': fetch_single_value(cursor, 'SELECT NumeLocalitate from tblLocalitati WHERE IDLocalitate = (SELECT LocalitateLucrare FROM tblLucrari WHERE IDLucrare = ?)', (id_lucrare,)),
        'adresa': fetch_single_value(cursor, 'SELECT AdresaLucrare FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet FROM tblJudete WHERE ID_Judet = (SELECT JudetLucrare FROM tblLucrari WHERE IDLucrare = ?)', (id_lucrare,)),

        'IDFirmaProiectare': fetch_single_value(cursor, 'SELECT IDFirmaProiectare FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'IDFirmaExecutie': fetch_single_value(cursor, 'SELECT IDFirmaExecutie FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'IDClient': fetch_single_value(cursor, 'SELECT IDClient FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'IDBeneficiar': fetch_single_value(cursor, 'SELECT IDBeneficiar FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        'IDPersoanaContact': fetch_single_value(cursor, 'SELECT IDPersoanaContact FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),

        # 'descrierea_proiectului': fetch_single_value(cursor, 'SELECT DescriereaProiectului FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'emitent_cu': fetch_single_value(cursor, 'SELECT DenumireInstitutie FROM tblEmitentCU WHERE IDEmitentCU = (SELECT EmitentCU FROM tblLucrari WHERE IDLucrare = ?)', (id_lucrare,)),

        # 'nr_cu': fetch_single_value(cursor, 'SELECT NumarCU FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'data_cu': fetch_single_value(cursor, 'SELECT DataCU FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),

        # 'IDIntocmit': fetch_single_value(cursor, 'SELECT IDIntocmit FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'IDVerificat': fetch_single_value(cursor, 'SELECT IDVerificat FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'facturare': fetch_single_value(cursor, 'SELECT FacturarePeClient FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleCU': fetch_single_value(cursor, 'SELECT CaleCU FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CalePlanIncadrare': fetch_single_value(cursor, 'SELECT CalePlanIncadrare FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CalePlanSituatie': fetch_single_value(cursor, 'SELECT CalePlanSituatie FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleMemoriuTehnic': fetch_single_value(cursor, 'SELECT CaleMemoriuTehnic FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleActeBeneficiar': fetch_single_value(cursor, 'SELECT CaleActeBeneficiar FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleActeFacturare': fetch_single_value(cursor, 'SELECT CaleActeFacturare FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleChitantaAPM': fetch_single_value(cursor, 'SELECT CaleChitantaAPM FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleChitantaDSP': fetch_single_value(cursor, 'SELECT CaleChitantaDSP FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'SuprafataMP': fetch_single_value(cursor, 'SELECT SuprafataMP FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'LungimeTraseuMetri': fetch_single_value(cursor, 'SELECT LungimeTraseuMetri FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleACCUConstructie': fetch_single_value(cursor, 'SELECT CaleACCUConstructie FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleAvizGiS': fetch_single_value(cursor, 'SELECT CaleAvizGiS FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleAvizCTEsauATR': fetch_single_value(cursor, 'SELECT CaleAvizCTEsauATR FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleExtraseCF': fetch_single_value(cursor, 'SELECT CaleExtraseCF FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CalePlanSituatiePDF': fetch_single_value(cursor, 'SELECT CalePlanSituatiePDF FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CalePlanSituatieDWG': fetch_single_value(cursor, 'SELECT CalePlanSituatieDWG FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
        # 'CaleRidicareTopoDWG': fetch_single_value(cursor, 'SELECT CaleRidicareTopoDWG FROM tblLucrari WHERE IDLucrare = ?', (id_lucrare,)),
    }


def get_Executie(cursor, id_executie):
    return {
        'nume': fetch_single_value(cursor, 'SELECT DenumireLucrare FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'localitate': fetch_single_value(cursor, 'SELECT LocalitateLucrare FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'adresa': fetch_single_value(cursor, 'SELECT AdresaLucrare FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'judet': fetch_single_value(cursor, 'SELECT NumeJudet FROM tblJudete WHERE ID_Judet = (SELECT JudetLucrare FROM tblExecutie WHERE IDExecutie = ?)', (id_executie,)),
        'IDFirmaProiectare': fetch_single_value(cursor, 'SELECT IDFirmaProiectare FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDFirmaExecutie': fetch_single_value(cursor, 'SELECT IDFirmaExecutie FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDClient': fetch_single_value(cursor, 'SELECT IDClient FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDBeneficiar': fetch_single_value(cursor, 'SELECT IDBeneficiar FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDUAT': fetch_single_value(cursor, 'SELECT IDEmitentAC FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDPersoanaContact': fetch_single_value(cursor, 'SELECT IDPersoanaContact FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),

        'numar_ac': fetch_single_value(cursor, 'SELECT NumarAC FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'data_ac': fetch_single_value(cursor, 'SELECT DataAC FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'valabilitate_ac': fetch_single_value(cursor, 'SELECT ValabilitateAC FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'valabilitate_executie': fetch_single_value(cursor, 'SELECT ValabilitateExecutie FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'data_incepere_executie': fetch_single_value(cursor, 'SELECT DataIncepereExecutie FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),

        'nr_cl': fetch_single_value(cursor, 'SELECT NrCL FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'nr_decizie_personal': fetch_single_value(cursor, 'SELECT NrDeciziePersonal FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'grafic_executie': fetch_single_value(cursor, 'SELECT GraficExecutie FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'data_incepere_grafic': fetch_single_value(cursor, 'SELECT DataIncepereGrafic FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'data_finalizare_grafic': fetch_single_value(cursor, 'SELECT DataFinalizareGrafic FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_sapare_sant': fetch_single_value(cursor, 'SELECT DPSapareSant FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_pozare_cablu': fetch_single_value(cursor, 'SELECT DPPozareCablu FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_montare_ptav': fetch_single_value(cursor, 'SELECT DPMontarePTAV FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_montare_stalpi': fetch_single_value(cursor, 'SELECT DPMontareStalpi FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_executie_foraj': fetch_single_value(cursor, 'SELECT DPExecutieForaj FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_montare_firide': fetch_single_value(cursor, 'SELECT DPMontareFiride FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'dp_stare_initiala': fetch_single_value(cursor, 'SELECT DPStareInitiala FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),


        'CaleACScanat': fetch_single_value(cursor, 'SELECT CaleACScanat FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CalePlanIncadrareACScanat': fetch_single_value(cursor, 'SELECT CalePlanIncadrareACScanat FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CalePlanSituatieACScanat': fetch_single_value(cursor, 'SELECT CalePlanSituatieACScanat FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleMemoriuTehnicACScanat': fetch_single_value(cursor, 'SELECT CaleMemoriuTehnicACScanat FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),


        'CalePlanIncadrarePTH': fetch_single_value(cursor, 'SELECT CalePlanIncadrarePTH FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CalePlanSituatiePTH': fetch_single_value(cursor, 'SELECT CalePlanSituatiePTH FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleSchemaMonofilaraJT': fetch_single_value(cursor, 'SELECT CaleSchemaMonofilaraJT FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleSchemaMonofilaraMT': fetch_single_value(cursor, 'SELECT CaleSchemaMonofilaraMT FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),

        'CaleInstruireColectiva': fetch_single_value(cursor, 'SELECT CaleInstruireColectiva FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleContractCitadin': fetch_single_value(cursor, 'SELECT CaleContractCitadin FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleContractSpatiiVerzi': fetch_single_value(cursor, 'SELECT CaleContractSpatiiVerzi FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleContractForaj': fetch_single_value(cursor, 'SELECT CaleContractForaj FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleAvizATR': fetch_single_value(cursor, 'SELECT CaleAvizATR FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleAvizCTE': fetch_single_value(cursor, 'SELECT CaleAvizCTE FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleContractRacordare': fetch_single_value(cursor, 'SELECT CaleContractRacordare FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'CaleContractExecutie': fetch_single_value(cursor, 'SELECT CaleContractExecutie FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'ContractExecutieDELGAZ': fetch_single_value(cursor, 'SELECT ContractExecutieDELGAZ FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'ContractExecutieCLIENT': fetch_single_value(cursor, 'SELECT ContractExecutieCLIENT FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),

        'IDDiriginteSantier': fetch_single_value(cursor, 'SELECT DiriginteSantier FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDRTEElectric': fetch_single_value(cursor, 'SELECT RTEElectric FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDRTEConstructii': fetch_single_value(cursor, 'SELECT RTEConstructii FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDResponsabilSSM': fetch_single_value(cursor, 'SELECT ResponsabilSSM FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDManagerProiectDELGAZ': fetch_single_value(cursor, 'SELECT ManagerProiectDELGAZ FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),
        'IDPersoanaContact': fetch_single_value(cursor, 'SELECT IDPersoanaContact FROM tblExecutie WHERE IDExecutie = ?', (id_executie,)),

    }


def get_Contact(cursor, id_contact):
    return {
        'nume': fetch_single_value(cursor, 'SELECT Nume FROM tblAngajati WHERE IDAngajat = ?', (id_contact,)),
        'telefon': fetch_single_value(cursor, 'SELECT Telefon FROM tblAngajati WHERE IDAngajat = ?', (id_contact,)),
    }


def create_document(model_path, context, final_destination, stampila_path=None):
    doc = DocxTemplate(model_path)
    # Try both strategies: replace an embedded picture placeholder and/or use InlineImage via a {{stampila}} variable
    if stampila_path:
        # 1) Attempt to replace an embedded placeholder image if present in the template
        try:
            doc.replace_pic("Placeholder_1.png", stampila_path)
        except Exception:
            # If the placeholder image doesn't exist, ignore and rely on InlineImage
            pass
        # 2) Also provide InlineImage in context for templates using {{stampila}}
        try:
            context = dict(context)  # avoid mutating caller's dict
            context.setdefault('stampila', InlineImage(doc, stampila_path, width=Mm(40)))
        except Exception:
            pass
    doc.render(context)
    nume, _ = os.path.splitext(os.path.basename(model_path))
    path_doc = os.path.join(final_destination, f'{nume}.docx')
    doc.save(path_doc)
    # Small delay to avoid file syncing/locking races on network drives
    time.sleep(1)
    cerere_pdf_path = convert_to_pdf(path_doc)
    if os.path.exists(path_doc):
        os.remove(path_doc)
    return cerere_pdf_path




def create_email(model_path, context, final_destination):
    doc = DocxTemplate(model_path)
    doc.render(context)
    path_doc = os.path.join(final_destination, 'Email.docx')
    doc.save(path_doc)


def merge_pdfs(pdf_list, output_path):
    merger = PdfMerger()
    for pdf in pdf_list:
        merger.append(pdf)
    merger.write(output_path)
    merger.close()


def merge_pdfs_print(pdf_list, output_path):
    merger = PdfMerger()
    for pdf in pdf_list:
        merger.append(pdf)
        x = count_pages(pdf)
        if x % 2 == 1:
            merger.append(pagina_goala)
    merger.write(output_path)
    merger.close()


def get_data(path_final, director_final, id_lucrare):
    final_destination = os.path.join(path_final, director_final)
    os.makedirs(final_destination, exist_ok=True,)
    conn = get_db_connection()
    cursor = conn.cursor()
    astazi = get_today_date()

    lucrare = get_Lucrare(cursor, id_lucrare)
    tblCU = get_CU(cursor, id_lucrare)
    EmitentCU = get_EmitentCU(cursor, tblCU['EmitentCU'])

    firma_proiectare = get_Firma_proiectare(
        cursor, lucrare['IDFirmaProiectare'])
    client = get_Client(cursor, lucrare['IDClient'])
    beneficiar = get_Beneficiar(cursor, lucrare['IDBeneficiar'])
    contact = get_Contact(cursor, lucrare['IDPersoanaContact'])

    intocmit = fetch_single_value(
        cursor, 'SELECT Nume FROM tblAngajati WHERE IDAngajat = ?', (tblCU['IDIntocmit'],))
    verificat = fetch_single_value(
        cursor, 'SELECT Nume FROM tblAngajati WHERE IDAngajat = ?', (tblCU['IDVerificat'],))

    return {
        'astazi': astazi,
        'lucrare': lucrare,
        'tblCU': tblCU,
        'firma_proiectare': firma_proiectare,
        'client': client,
        'beneficiar': beneficiar,
        'contact': contact,
        'final_destination': final_destination,
        'intocmit': intocmit,
        'verificat': verificat,
        'EmitentCU': EmitentCU,
    }


def get_data_executie(path_final, director_final, id_lucrare):
    final_destination = os.path.join(path_final, director_final)
    os.makedirs(final_destination, exist_ok=True,)
    conn = get_db_connection()
    cursor = conn.cursor()
    astazi = get_today_date()

    lucrare = get_Lucrare(cursor, id_lucrare)
    tblCU = get_CU(cursor, id_lucrare)
    tblIncepereExecutie = get_IncepereExecutie(cursor, id_lucrare)

    EmitentAC = get_EmitentCU(cursor, tblCU['EmitentCU'])

    firma_proiectare = get_Firma_proiectare(cursor, lucrare['IDFirmaProiectare'])
    firma_executie = get_Firma_executie(cursor, lucrare['IDFirmaExecutie'])
    client = get_Client(cursor, lucrare['IDClient'])
    beneficiar = get_Beneficiar(cursor, lucrare['IDBeneficiar'])
    contact = get_Contact(cursor, lucrare['IDPersoanaContact'])

    diriginte_santier = get_Diriginte_Santier(cursor, tblIncepereExecutie['ID_DiriginteSantier'])
    manager_proiect = get_Manager_Proiect(cursor, tblIncepereExecutie['ID_ManagerProiect'])
    rte = get_RTE(cursor, tblIncepereExecutie['ID_RTE'])
    rte_constructii = get_RTE_constructii(cursor, tblIncepereExecutie['ID_RTEConstructii'])
    responsabil_ssm = get_responsabil_SSM(cursor, tblIncepereExecutie['ID_SSM'])

    return {
        'astazi': astazi,
        'lucrare': lucrare,
        'tblIncepereExecutie': tblIncepereExecutie,
        'tblCU': tblCU,

        'EmitentAC': EmitentAC,

        'firma_proiectare': firma_proiectare,
        'firma_executie': firma_executie,
        'client': client,
        'beneficiar': beneficiar,

        'rte': rte,
        'rte_constructii': rte_constructii,
        'manager_proiect': manager_proiect,
        'diriginte_santier': diriginte_santier,
        'responsabil_ssm': responsabil_ssm,

        'final_destination': final_destination,
        'contact': contact,
    }


def get_data_finalizare(path_final, director_final, id_lucrare):
    final_destination = os.path.join(path_final, director_final)
    os.makedirs(final_destination, exist_ok=True,)
    conn = get_db_connection()
    cursor = conn.cursor()
    astazi = get_today_date()

    lucrare = get_Lucrare(cursor, id_lucrare)
    tblCU = get_CU(cursor, id_lucrare)
    tblIncepereExecutie = get_IncepereExecutie(cursor, id_lucrare)
    tblFinalizare = get_Finalizare(cursor, id_lucrare)

    EmitentAC = get_EmitentCU(cursor, tblCU['EmitentCU'])

    firma_proiectare = get_Firma_proiectare(cursor, lucrare['IDFirmaProiectare'])
    firma_executie = get_Firma_executie(cursor, lucrare['IDFirmaExecutie'])
    client = get_Client(cursor, lucrare['IDClient'])
    beneficiar = get_Beneficiar(cursor, lucrare['IDBeneficiar'])
    contact = get_Contact(cursor, lucrare['IDPersoanaContact'])

    diriginte_santier = get_Diriginte_Santier(cursor, tblIncepereExecutie['ID_DiriginteSantier'])
    manager_proiect = get_Manager_Proiect(cursor, tblIncepereExecutie['ID_ManagerProiect'])
    rte = get_RTE(cursor, tblIncepereExecutie['ID_RTE'])
    rte_constructii = get_RTE_constructii(cursor, tblIncepereExecutie['ID_RTEConstructii'])
    responsabil_ssm = get_responsabil_SSM(cursor, tblIncepereExecutie['ID_SSM'])

    return {
        'astazi': astazi,
        'lucrare': lucrare,
        'tblIncepereExecutie': tblIncepereExecutie,
        'tblCU': tblCU,
        'tblFinalizare': tblFinalizare,

        'EmitentAC': EmitentAC,

        'firma_proiectare': firma_proiectare,
        'firma_executie': firma_executie,
        'client': client,
        'beneficiar': beneficiar,

        'rte': rte,
        'rte_constructii': rte_constructii,
        'manager_proiect': manager_proiect,
        'diriginte_santier': diriginte_santier,
        'responsabil_ssm': responsabil_ssm,

        'final_destination': final_destination,
        'contact': contact,
    }


def facturare(id_lucrare):
    conn = get_db_connection()
    cursor = conn.cursor()
    tblCU = get_CU(cursor, id_lucrare)
    lucrare = get_Lucrare(cursor, id_lucrare)
    firma_proiectare = get_Firma_proiectare(
        cursor, lucrare['IDFirmaProiectare'])
    client = get_Client(cursor, lucrare['IDClient'])
    if tblCU['FacturarePeClient'] == False:
        return {
            'firma_facturare': firma_proiectare['nume'],
            'cui_firma_facturare': firma_proiectare['CUI'],
            'nr_reg_com_facturare': firma_proiectare['NrRegCom'],
            'localitate_facturare': firma_proiectare['localitate'],
            'adresa_facturare': firma_proiectare['adresa'],
            'judet_facturare': firma_proiectare['judet'],
        }
    else:
        return {
            'firma_facturare': client['nume'],
            'cui_firma_facturare': client['CUI'],
            'nr_reg_com_facturare': client['NrRegCom'],
            'localitate_facturare': client['localitate'],
            'adresa_facturare': client['adresa'],
            'judet_facturare': client['judet'],
        }


def copy_file(file_path, path_final, director_final, file_name: str):
    file = file_path.strip('"')
    shutil.copy(file, os.path.join(path_final, director_final, file_name))


def move_file(file_path, path_final, director_final, file_name: str):
    file = file_path.strip('"')
    shutil.move(file, os.path.join(path_final, director_final, file_name))


def copy_file_prefix(file_path, path_final, director_final, prefix=None):
    file = file_path.strip('"')
    filename = os.path.basename(file)
    new_filename = prefix + filename
    shutil.copy(file, os.path.join(path_final, director_final, new_filename))


def count_pages_ISU(cerere_path, cu_path, plan_incadrare_path, plan_situatie_path, cale_memoriu, cale_acte):
    cerere = count_pages(cerere_path)
    cu = count_pages(cu_path)
    plan_incadrare = count_pages(plan_incadrare_path)
    plan_situatie = count_pages(plan_situatie_path)
    memoriu = count_pages(cale_memoriu)
    acte = count_pages(cale_acte)
    return {
        'cerere': cerere,
        'cu': cu,
        'plan_incadrare': plan_incadrare,
        'plan_situatie': plan_situatie,
        'memoriu_tehnic': memoriu,
        'acte_facturare': acte
    }


def get_date(field_entry):
    if field_entry and isinstance(field_entry, str):
        date_str = field_entry
        date_obj = dt.strptime(date_str, '%d-%m-%Y')
    else:
        date_obj = field_entry
    return date_obj.strftime('%d-%m-%Y')


def get_month(entry):
    luna = entry.month
    if luna == 1:
        return "Ianuarie"
    if luna == 2:
        return "Februarie"
    if luna == 3:
        return "Martie"
    if luna == 4:
        return "Aprilie"
    if luna == 5:
        return "Mai"
    if luna == 6:
        return "Iunie"
    if luna == 7:
        return "Iulie"
    if luna == 8:
        return "August"
    if luna == 9:
        return "Septembrie"
    if luna == 10:
        return "Octombrie"
    if luna == 11:
        return "Noiembrie"
    if luna == 12:
        return "Decembrie"


def get_year(entry):
    return entry.year


def get_day(entry):
    return entry.day


def genereaza_grafic_executie(context, final_destination):
    wb = openpyxl.load_workbook(context['file_path'])
    ws = wb.active

    # Load the image file

    img = Image(context['logo_path'])
    img.anchor = 'A1'
    # Add the image to the worksheet
    ws.add_image(img)

    cm_to_pixels = 37.795275591
    desired_width_cm = 5
    desired_height_cm = 3.5
    desired_width_px = desired_width_cm * cm_to_pixels
    desired_height_px = desired_height_cm * cm_to_pixels

    img2 = Image(context['stampila_path'])
    img2.anchor = 'D19'
    img2.width = desired_width_px
    img2.height = desired_height_px
    # Add the image to the worksheet
    ws.add_image(img2)

    # adaugam emitentul AC
    vizat = f'VIZAT {context['emitent_ac']}'

    ws['D2'] = vizat.upper()

    ws['B5'] = f"{context['nume_lucrare']} din {context['localitate_lucrare']}, {context['adresa_lucrare']}, județ {context['judet_lucrare']}"

    ws['B6'] = f"Beneficiar: {context['nume_client']}"

    an_incepere = f'Anul {context['an_lucrare']}'
    ws['C8'] = an_incepere

    if context['luna_incepere'] == context['luna_finalizare']:
        ws.merge_cells('C10:D10')
        ws['C10'] = context['luna_incepere']
    else:
        ws['C10'] = context['luna_incepere']
        ws['D10'] = context['luna_finalizare']

    incepere = f'de la {context["data_incepere"]}'
    ws['C11'] = incepere
    finalizare = f'până la {context["data_finalizare"]}'
    ws['D11'] = finalizare

    total_height_cm = 7  # Total height for the rows (in cm)

    # Define the starting and ending row for the data in column A (A12 to A18)
    start_row = 12
    end_row = 18

    # Determine the number of rows to be used, dynamically between 3 and 7
    # You can set this dynamically or for testing purposes, let's assume we want 5 rows

    lista_valori = []

    if context['sapare_sant']:
        lista_valori.append('SĂPARE MANUALĂ ȘANȚ - Domeniu PUBLIC')
    if context['pozare_cablu']:
        lista_valori.append(
            'POZARE CABLU ȘI ACOPERIRE ȘANȚ CU PĂMÂNT - Domeniu PUBLIC')
    if context['montare_ptav']:
        lista_valori.append('MONTARE PTAv - Domeniu PUBLIC')
    if context['montare_stalpi']:
        lista_valori.append('MONTARE STÂLPI - Domeniul PUBLIC')
    if context['executie_foraj']:
        lista_valori.append('EXECUȚIE FORAJ - Domeniul PUBLIC')
    if context['montare_firide']:
        lista_valori.append('MONTARE FIRIDE - Domeniul PUBLIC')
    if context['aducere_stare_initiala']:
        lista_valori.append(
            'ADUCERE TEREN LA STAREA INIȚIALĂ - Domeniul PUBLIC')

    # Change this value to vary the number of rows
    num_rows = len(lista_valori)

    # Ensure num_rows is between the minimum and maximum values
    # num_rows = max(min_rows, min(num_rows, max_rows))

    # Calculate the row height (in Excel, 1 row height = approximately 0.75 points, 1 cm ≈ 28.35 points)
    row_height_points = total_height_cm * 28.35 / num_rows

    # Set row heights from row A12 to A18)
    for row in range(start_row, start_row + num_rows):
        ws.row_dimensions[row].height = row_height_points

    # Dynamically reference columns using get_column_letter (e.g., "A" and "B")
    col_A = get_column_letter(1)  # Column A
    col_B = get_column_letter(2)  # Column B

    # Number the rows in column A
    for i in range(num_rows):
        # Fill column A with row numbers from 1 to num_rows
        ws[f"{col_A}{start_row + i}"] = i + 1

    # Insert data into column B
    for i in range(num_rows):
        # Example data, change as needed
        ws[f"{col_B}{start_row + i}"] = lista_valori[i]

    # Define a border style (thin borders for all sides)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply the border to all cells in the table (A12:B18)
    for row in range(start_row, start_row + num_rows):
        for col in range(1, 2):  # Column A and B
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    for row in range(start_row, start_row + num_rows):
        for col in range(2, 3):  # Column A and B
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True)

    gray_fill = PatternFill(start_color="D3D3D3",
                            end_color="D3D3D3", fill_type="solid")
    for row in range(start_row, start_row + num_rows):
        for col in range(3, 5):  # Columns C (3) and D (4)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Apply the border
            cell.fill = gray_fill  # Apply the gray fill color
            cell.value = ""  # Leave the cell empty (no data)

    for row in range(start_row, 19):
        cell = ws.cell(row=row, column=1)  # Column A

        # Check if the cell is truly empty or contains only invisible characters
        if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
            # If the row is empty (including spaces or invisible characters), set its height to 0
            ws.row_dimensions[row].height = 0.1
        else:
            ws.row_dimensions[row].height = row_height_points

    # Save the workbook
    wb.save(final_destination)


def genereaza_grafic_executie_iasi(context, final_destination):
    wb = openpyxl.load_workbook(context['file_path'])
    ws = wb.active

    # Load the image file

    img = Image(context['logo_path'])
    img.anchor = 'A1'
    # Add the image to the worksheet
    ws.add_image(img)

    cm_to_pixels = 37.795275591
    desired_width_cm = 5
    desired_height_cm = 3.5
    desired_width_px = desired_width_cm * cm_to_pixels
    desired_height_px = desired_height_cm * cm_to_pixels

    img2 = Image(context['stampila_path'])
    img2.anchor = 'E19'
    img2.width = desired_width_px
    img2.height = desired_height_px
    # Add the image to the worksheet
    ws.add_image(img2)

    # adaugam emitentul AC
    vizat = f'VIZAT {context['emitent_ac']}'
    ws['E2'] = vizat.upper()

    ws['B5'] = f"{context['nume_lucrare']} din {context['localitate_lucrare']}, {context['adresa_lucrare']}, județ {context['judet_lucrare']}"

    ws['B6'] = f"Beneficiar: {context['nume_client']}"

    an_incepere = f'Anul {context['an_lucrare']}'
    ws['C8'] = an_incepere

    if context['luna_incepere'] == context['luna_finalizare']:
        ws.merge_cells('C10:D10')
        ws['C10'] = context['luna_incepere']
    else:
        ws['C10'] = context['luna_incepere']
        ws['D10'] = context['luna_finalizare']

    incepere = f'de la {context["data_incepere"]}'
    ws['C11'] = incepere
    finalizare = f'până la {context["data_finalizare"]}'
    ws['D11'] = finalizare

    total_height_cm = 7  # Total height for the rows (in cm)

    # Define the starting and ending row for the data in column A (A12 to A18)
    start_row = 12
    end_row = 18

    # Determine the number of rows to be used, dynamically between 3 and 7
    # You can set this dynamically or for testing purposes, let's assume we want 5 rows

    lista_valori = []

    if context['sapare_sant']:
        lista_valori.append('SĂPARE MANUALĂ ȘANȚ - Domeniu PUBLIC')
    if context['pozare_cablu']:
        lista_valori.append(
            'POZARE CABLU ȘI ACOPERIRE ȘANȚ CU PĂMÂNT - Domeniu PUBLIC')
    if context['montare_ptav']:
        lista_valori.append('MONTARE PTAv - Domeniu PUBLIC')
    if context['montare_stalpi']:
        lista_valori.append('MONTARE STÂLPI - Domeniul PUBLIC')
    if context['executie_foraj']:
        lista_valori.append('EXECUȚIE FORAJ - Domeniul PUBLIC')
    if context['montare_firide']:
        lista_valori.append('MONTARE FIRIDE - Domeniul PUBLIC')
    if context['aducere_stare_initiala']:
        lista_valori.append(
            'ADUCERE TEREN LA STAREA INIȚIALĂ - Domeniul PUBLIC')

    # Change this value to vary the number of rows
    num_rows = len(lista_valori)

    # Ensure num_rows is between the minimum and maximum values
    # num_rows = max(min_rows, min(num_rows, max_rows))

    # Calculate the row height (in Excel, 1 row height = approximately 0.75 points, 1 cm ≈ 28.35 points)
    row_height_points = total_height_cm * 28.35 / num_rows

    # Set row heights from row A12 to A18)
    for row in range(start_row, start_row + num_rows):
        ws.row_dimensions[row].height = row_height_points

    # Dynamically reference columns using get_column_letter (e.g., "A" and "B")
    col_A = get_column_letter(1)  # Column A
    col_B = get_column_letter(2)  # Column B

    # Number the rows in column A
    for i in range(num_rows):
        # Fill column A with row numbers from 1 to num_rows
        ws[f"{col_A}{start_row + i}"] = i + 1

    # Insert data into column B
    for i in range(num_rows):
        # Example data, change as needed
        ws[f"{col_B}{start_row + i}"] = lista_valori[i]

    # Define a border style (thin borders for all sides)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply the border to all cells in the table (A12:B18)
    for row in range(start_row, start_row + num_rows):
        for col in range(1, 2):  # Column A and B
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    for row in range(start_row, start_row + num_rows):
        for col in range(2, 3):  # Column A and B
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True)

    gray_fill = PatternFill(start_color="D3D3D3",
                            end_color="D3D3D3", fill_type="solid")
    for row in range(start_row, start_row + num_rows):
        for col in range(3, 5):  # Columns C (3) and D (4)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Apply the border
            cell.fill = gray_fill  # Apply the gray fill color
            cell.value = ""  # Leave the cell empty (no data)

    for row in range(start_row, 19):
        cell = ws.cell(row=row, column=1)  # Column A

        # Check if the cell is truly empty or contains only invisible characters
        if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
            # If the row is empty (including spaces or invisible characters), set its height to 0
            ws.row_dimensions[row].height = 0.1
        else:
            ws.row_dimensions[row].height = row_height_points

    # Save the workbook
    wb.save(final_destination)


def aduna_luni(date_obj, luni: int):
    new_date = date_obj + relativedelta(months=luni)
    formatted_date = new_date.strftime('%d-%m-%Y')
    return formatted_date


def custom_round(value_1):
    value = value_1 * 0.01
    return math.ceil(value) if value % 1 >= 0.5 else math.floor(value)

def diferenta_taxa(taxa_ac, taxa_reala):
    # If taxa AC is greater or equal, no difference is owed
    if round(taxa_ac) >= round(taxa_reala):
        return "-"
    diferenta = taxa_reala - taxa_ac
    return f"{diferenta:.2f}"

def calculeaza_taxa_reala(valoare_reala, valoare_ac):
    # Afișează taxa doar dacă valoarea reală este mai mare decât valoarea AC; altfel "-"
    if round(valoare_reala) <= round(valoare_ac):
        return "-"
    taxa_rotunjita = custom_round(valoare_reala)
    return f"{taxa_rotunjita:.2f}"


def recreate_pdf(input_path, output_path):
    """
    Recreează un fișier PDF pentru a elimina semnătura electronică.
    """
    reader = PdfReader(input_path)
    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    with open(output_path, "wb") as output_file:
        writer.write(output_file)

def merge_pdfs_signed(pdf_list, output_path):
    """
    Combină fișierele PDF, recreând fișierele semnate electronic dacă este necesar.
    """
    merger = PdfMerger()

    for pdf in pdf_list:
        try:
            # Verificăm dacă fișierul este semnat sau are restricții
            reader = PdfReader(pdf)
            if reader.is_encrypted:
                print(f"Fișierul {pdf} este semnat electronic. Se recreează...")
                temp_pdf = pdf.replace(".pdf", "_recreated.pdf")
                recreate_pdf(pdf, temp_pdf)
                merger.append(temp_pdf)
                # Ștergem fișierul temporar după utilizare
                os.remove(temp_pdf)
            else:
                merger.append(pdf)
        except Exception as e:
            print(f"Eroare la procesarea fișierului {pdf}: {e}")
    merger.write(output_path)
    merger.close()